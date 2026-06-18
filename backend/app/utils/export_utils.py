"""Export utility for generating CSV, Excel, and PDF files"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportGenerator:
    """Generate exports in CSV, Excel, and PDF formats"""

    @staticmethod
    def generate_csv(data: List[Dict[str, Any]], headers: List[str] = None) -> bytes:
        """
        Generate CSV data from a list of dictionaries

        Args:
            data: List of dictionaries containing the data
            headers: List of column headers (if None, uses dict keys from first row)

        Returns:
            CSV content as bytes
        """
        if not data:
            return b''

        if headers is None:
            headers = list(data[0].keys()) if data else []

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue().encode('utf-8')

    @staticmethod
    def generate_excel(
        data: List[Dict[str, Any]],
        headers: List[str] = None,
        title: str = "Export",
        sheet_name: str = "Data"
    ) -> bytes:
        """
        Generate Excel file from a list of dictionaries

        Args:
            data: List of dictionaries containing the data
            headers: List of column headers
            title: Title of the export
            sheet_name: Name of the worksheet

        Returns:
            Excel file content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not data:
            return b''

        if headers is None:
            headers = list(data[0].keys()) if data else []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:' + chr(64 + len(headers)) + '1')

        # Add export timestamp
        ws['A2'] = f"Exported on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
        ws['A2'].font = Font(italic=True, size=10)

        # Add headers
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(header, '')
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_pdf(
        data: List[Dict[str, Any]],
        headers: List[str] = None,
        title: str = "Export",
        company_name: str = "Company"
    ) -> bytes:
        """
        Generate PDF file from a list of dictionaries

        Args:
            data: List of dictionaries containing the data
            headers: List of column headers
            title: Title of the export
            company_name: Company name for header

        Returns:
            PDF file content as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        if not data:
            return b''

        if headers is None:
            headers = list(data[0].keys()) if data else []

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=6,
            alignment=1  # Center
        )
        story.append(Paragraph(title, title_style))

        # Add company and timestamp
        meta_style = ParagraphStyle(
            'Meta',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=1
        )
        story.append(Paragraph(
            f"Company: {company_name} | Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            meta_style
        ))

        # Prepare table data
        table_data = [headers]

        # Convert data to table format
        for row in data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                # Convert to string and truncate if necessary
                str_value = str(value)[:50]  # Limit cell content to 50 chars
                table_row.append(str_value)
            table_data.append(table_row)

        # Create table
        if table_data:
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ]))
            story.append(table)

        # Build PDF
        doc.build(story)
        return output.getvalue()

    @staticmethod
    def get_filename(data_type: str, format_type: str) -> str:
        """
        Generate a filename for the export

        Args:
            data_type: Type of data being exported
            format_type: Format of the export (csv, excel, pdf)

        Returns:
            Filename with timestamp
        """
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        extension_map = {
            'csv': 'csv',
            'excel': 'xlsx',
            'pdf': 'pdf'
        }
        extension = extension_map.get(format_type, format_type)
        return f"{data_type}_export_{timestamp}.{extension}"
